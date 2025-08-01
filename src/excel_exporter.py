"""
Excel Export functionality for Git Analytics Data
"""

import logging
from typing import Dict, List, Any, Optional
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

from .analytics_engine import AnalyticsResult
from .config import get_config


logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel export functionality for analytics data"""
    
    def __init__(self):
        """Initialize Excel exporter"""
        self.config = get_config()
        
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.sub_header_font = Font(bold=True, size=11)
        self.sub_header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    async def export_analytics(self, analytics_results: Dict[str, AnalyticsResult], 
                              output_filename: Optional[str] = None) -> Path:
        """Export analytics results to Excel file with multiple sheets"""
        
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") if self.config.output.include_timestamp else ""
            suffix = f"_{timestamp}" if timestamp else ""
            output_filename = f"{self.config.output.filename_prefix}{suffix}.xlsx"
        
        output_path = self.config.get_output_path(output_filename)
        
        logger.info(f"Exporting analytics to Excel: {output_path}")
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create summary sheet
        await self._create_summary_sheet(wb, analytics_results)
        
        # Create detailed sheets for each repository
        for repo_key, result in analytics_results.items():
            await self._create_repository_sheet(wb, repo_key, result)
        
        # Create consolidated analysis sheets
        await self._create_consolidated_sheets(wb, analytics_results)
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel export completed: {output_path}")
        
        return output_path
    
    async def _create_summary_sheet(self, wb: Workbook, analytics_results: Dict[str, AnalyticsResult]):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary Dashboard", 0)
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "Git Repository Analytics Dashboard"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = self.center_alignment
        
        # Generation info
        ws['A3'] = "Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = "Total Repositories:"
        ws['B4'] = len(analytics_results)
        
        # Summary statistics
        row = 6
        ws[f'A{row}'] = "Repository Summary"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        
        # Headers
        row += 1
        headers = ["Repository", "Total Commits", "Authors", "Branches", "Pull Requests", "Latest Activity"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data rows
        for repo_key, result in analytics_results.items():
            row += 1
            commit_analytics = result.commit_analytics
            author_analytics = result.author_analytics
            branch_analytics = result.branch_analytics
            pr_analytics = result.pull_request_analytics
            
            data = [
                repo_key,
                commit_analytics.get('total_commits', 0),
                author_analytics.get('total_authors', 0),
                branch_analytics.get('total_branches', 0),
                pr_analytics.get('total_pull_requests', 0),
                commit_analytics.get('last_commit_date', 'N/A')[:10] if commit_analytics.get('last_commit_date') else 'N/A'
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1 and col < 6:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _create_repository_sheet(self, wb: Workbook, repo_key: str, result: AnalyticsResult):
        """Create detailed sheet for a single repository"""
        # Sanitize sheet name (Excel has limitations)
        sheet_name = repo_key.replace('/', '_')[:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)
        
        row = 1
        
        # Repository header
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = f"Repository Analysis: {repo_key}"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].alignment = self.center_alignment
        row += 2
        
        # Repository info
        repo_info = result.repository_info
        ws[f'A{row}'] = "Repository Information"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        row += 1
        
        info_data = [
            ("Project", repo_info.get('project', 'N/A')),
            ("Repository", repo_info.get('repository', 'N/A')),
            ("Default Branch", repo_info.get('default_branch', 'N/A')),
            ("Size", repo_info.get('size', 'N/A')),
            ("Analysis Date", repo_info.get('analysis_date', 'N/A')[:19] if repo_info.get('analysis_date') else 'N/A')
        ]
        
        for key, value in info_data:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Commit Analytics
        if result.commit_analytics:
            row = await self._add_commit_analytics_section(ws, row, result.commit_analytics)
        
        # Author Analytics
        if result.author_analytics:
            row = await self._add_author_analytics_section(ws, row, result.author_analytics)
        
        # Time Analytics
        if result.time_analytics:
            row = await self._add_time_analytics_section(ws, row, result.time_analytics)
        
        # Code Quality
        if result.code_quality:
            row = await self._add_code_quality_section(ws, row, result.code_quality)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _add_commit_analytics_section(self, ws, start_row: int, commit_analytics: Dict[str, Any]) -> int:
        """Add commit analytics section to worksheet"""
        row = start_row
        
        # Section header
        ws[f'A{row}'] = "Commit Analytics"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        row += 1
        
        # Basic metrics
        metrics = [
            ("Total Commits", commit_analytics.get('total_commits', 0)),
            ("Total Additions", commit_analytics.get('total_additions', 0)),
            ("Total Edits", commit_analytics.get('total_edits', 0)),
            ("Total Deletions", commit_analytics.get('total_deletions', 0)),
            ("Merge Commits", commit_analytics.get('merge_commits', 0)),
            ("Regular Commits", commit_analytics.get('regular_commits', 0)),
            ("Merge Ratio", f"{commit_analytics.get('merge_ratio', 0):.2%}"),
            ("Avg Message Length", f"{commit_analytics.get('average_message_length', 0):.1f}"),
            ("First Commit", commit_analytics.get('first_commit_date', 'N/A')[:10] if commit_analytics.get('first_commit_date') else 'N/A'),
            ("Last Commit", commit_analytics.get('last_commit_date', 'N/A')[:10] if commit_analytics.get('last_commit_date') else 'N/A')
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # Commits by day of week
        commits_by_day = commit_analytics.get('commits_by_day_of_week', {})
        if commits_by_day:
            row += 1
            ws[f'A{row}'] = "Commits by Day of Week"
            ws[f'A{row}'].font = self.sub_header_font
            row += 1
            
            for day, count in commits_by_day.items():
                ws[f'A{row}'] = day
                ws[f'B{row}'] = count
                row += 1
        
        return row + 2
    
    async def _add_author_analytics_section(self, ws, start_row: int, author_analytics: Dict[str, Any]) -> int:
        """Add author analytics section to worksheet"""
        row = start_row
        
        # Section header
        ws[f'A{row}'] = "Author Analytics"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        row += 1
        
        # Basic metrics
        ws[f'A{row}'] = "Total Authors"
        ws[f'B{row}'] = author_analytics.get('total_authors', 0)
        row += 1
        
        ws[f'A{row}'] = "Bus Factor (50%)"
        ws[f'B{row}'] = author_analytics.get('bus_factor_50_percent', 0)
        row += 1
        
        ws[f'A{row}'] = "Bus Factor (80%)"
        ws[f'B{row}'] = author_analytics.get('bus_factor_80_percent', 0)
        row += 1
        
        # Top contributors by commits
        top_contributors = author_analytics.get('top_contributors_by_commits', {})
        if top_contributors:
            row += 1
            ws[f'A{row}'] = "Top Contributors by Commits"
            ws[f'A{row}'].font = self.sub_header_font
            row += 1
            
            # Headers
            ws[f'A{row}'] = "Author"
            ws[f'B{row}'] = "Commits"
            ws[f'C{row}'] = "Total Changes"
            for col in ['A', 'B', 'C']:
                ws[f'{col}{row}'].font = self.header_font
                ws[f'{col}{row}'].fill = self.header_fill
            row += 1
            
            for author, stats in list(top_contributors.items())[:10]:  # Top 10
                ws[f'A{row}'] = author
                ws[f'B{row}'] = stats.get('commits', 0)
                ws[f'C{row}'] = stats.get('total_changes', 0)
                row += 1
        
        return row + 2
    
    async def _add_time_analytics_section(self, ws, start_row: int, time_analytics: Dict[str, Any]) -> int:
        """Add time analytics section to worksheet"""
        row = start_row
        
        # Section header
        ws[f'A{row}'] = "Time Analytics"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        row += 1
        
        # Work patterns
        metrics = [
            ("Weekend Commits", time_analytics.get('weekend_commits', 0)),
            ("Weekday Commits", time_analytics.get('weekday_commits', 0)),
            ("Weekend Ratio", f"{time_analytics.get('weekend_ratio', 0):.2%}"),
            ("After Hours Commits", time_analytics.get('after_hours_commits', 0)),
            ("After Hours Ratio", f"{time_analytics.get('after_hours_ratio', 0):.2%}")
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        # PR cycle times
        if 'average_pr_cycle_time_hours' in time_analytics:
            row += 1
            ws[f'A{row}'] = "Pull Request Metrics"
            ws[f'A{row}'].font = self.sub_header_font
            row += 1
            
            pr_metrics = [
                ("Avg PR Cycle Time (hours)", f"{time_analytics.get('average_pr_cycle_time_hours', 0):.1f}"),
                ("Median PR Cycle Time (hours)", f"{time_analytics.get('median_pr_cycle_time_hours', 0):.1f}"),
                ("Completed PRs", time_analytics.get('completed_prs', 0))
            ]
            
            for metric, value in pr_metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                row += 1
        
        return row + 2
    
    async def _add_code_quality_section(self, ws, start_row: int, code_quality: Dict[str, Any]) -> int:
        """Add code quality section to worksheet"""
        row = start_row
        
        # Section header
        ws[f'A{row}'] = "Code Quality Indicators"
        ws[f'A{row}'].font = self.sub_header_font
        ws[f'A{row}'].fill = self.sub_header_fill
        row += 1
        
        # Quality metrics
        metrics = [
            ("Total Additions", code_quality.get('total_additions', 0)),
            ("Total Deletions", code_quality.get('total_deletions', 0)),
            ("Delete/Add Ratio", f"{code_quality.get('delete_add_ratio', 0):.3f}"),
            ("Large Commits (>500 changes)", code_quality.get('large_commits_count', 0)),
            ("Refactoring Commits", code_quality.get('refactoring_commits_count', 0))
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            row += 1
        
        return row + 2
    
    async def _create_consolidated_sheets(self, wb: Workbook, analytics_results: Dict[str, AnalyticsResult]):
        """Create consolidated analysis sheets"""
        
        # Create commits comparison sheet
        await self._create_commits_comparison_sheet(wb, analytics_results)
        
        # Create authors comparison sheet
        await self._create_authors_comparison_sheet(wb, analytics_results)
    
    async def _create_commits_comparison_sheet(self, wb: Workbook, analytics_results: Dict[str, AnalyticsResult]):
        """Create sheet comparing commit metrics across repositories"""
        ws = wb.create_sheet("Commits Comparison")
        
        # Headers
        headers = ["Repository", "Total Commits", "Additions", "Deletions", "Edits", "Merge Ratio", "Authors"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        row = 2
        for repo_key, result in analytics_results.items():
            commit_analytics = result.commit_analytics
            author_analytics = result.author_analytics
            
            data = [
                repo_key,
                commit_analytics.get('total_commits', 0),
                commit_analytics.get('total_additions', 0),
                commit_analytics.get('total_deletions', 0),
                commit_analytics.get('total_edits', 0),
                f"{commit_analytics.get('merge_ratio', 0):.2%}",
                author_analytics.get('total_authors', 0)
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1 and col != 6:  # Numeric columns except merge ratio
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    async def _create_authors_comparison_sheet(self, wb: Workbook, analytics_results: Dict[str, AnalyticsResult]):
        """Create sheet comparing author metrics across repositories"""
        ws = wb.create_sheet("Authors Comparison")
        
        # Headers
        headers = ["Repository", "Total Authors", "Bus Factor 50%", "Bus Factor 80%", "Top Author Commits", "Top Author Changes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # Data
        row = 2
        for repo_key, result in analytics_results.items():
            author_analytics = result.author_analytics
            
            # Get top contributor stats
            top_contributors = author_analytics.get('top_contributors_by_commits', {})
            top_author_commits = 0
            top_author_changes = 0
            
            if top_contributors:
                first_author = next(iter(top_contributors.values()))
                top_author_commits = first_author.get('commits', 0)
                top_author_changes = first_author.get('total_changes', 0)
            
            data = [
                repo_key,
                author_analytics.get('total_authors', 0),
                author_analytics.get('bus_factor_50_percent', 0),
                author_analytics.get('bus_factor_80_percent', 0),
                top_author_commits,
                top_author_changes
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = self.border
                if col > 1:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width 